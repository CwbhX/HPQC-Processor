from openpyxl import load_workbook
from testscript import TestScript
import glob, os

class Loader:
	def __init__(self, directory, config):
		self.directory = directory
		self.config = config
		self.testScripts = []
	
	def getFiles(self): # Get all excel files in the directory passed to this Loader instance and their absolute paths returned in a list
		os.chdir(self.directory)
		files = []
		
		for spreadsheet in glob.glob("*.xlsx"):
			if self.directory.endswith("/"):
				files.append(self.directory + spreadsheet)
			else:
				files.append(self.directory + "/" + spreadsheet)
		return sorted(files) # Return the files in the proper sorted order!	
	
	
	def getCellsInSheet(self, spreadsheet):
		cells = []
		
		for row in range(1, spreadsheet.max_row+1):
			for column in range(1, spreadsheet.max_column+1):
				cells.append(spreadsheet.cell(row=row, column=column))
				
		return cells
	
	
	def getCellContaining(self, spreadsheet, text): # Get a the cell containing a keyword, returns first instance if multiple
		cells = self.getCellsInSheet(spreadsheet)
		
		for cell in cells:
			if text in str(cell.value):
				return cell
		
		return None # If we never find a cell with the contained text
	
	
	def getAdjacentCellTo(self, spreadsheet, cell): # Get the cell to the right of a cell
		if cell == None:
			return None;
		
		adjCell = spreadsheet.cell(row=cell.row, column=cell.col_idx+1)
		
		if adjCell.value == None:
			adjCell = spreadsheet.cell(row=cell.row, column=cell.col_idx+2)
			return adjCell
		else:
			return adjCell
	
	
	def getAdjacentCellForKey(self, spreadsheet, keyText): # Return a cell that is adjacent to a cell with a certain value as a key
		originCell = self.getCellContaining(spreadsheet, keyText)
		return self.getAdjacentCellTo(spreadsheet, originCell)
	
	
	def loadStepData(self, spreadsheet, stepHeaderCell): # load step data from spread sheet, return dictionary with key step and value a list of length 2 with step description and expected result
		stepData = {}
		
		stepCount = 1
		stepCell = spreadsheet.cell(row=stepHeaderCell.row + stepCount, column=stepHeaderCell.col_idx) # Get first step cell
		while type(stepCell.value) == int:
			#                                                            Step Details                                                 Expected Result
			stepData[stepCell.value] = [spreadsheet.cell(row=stepCell.row, column=stepCell.col_idx+1).value, spreadsheet.cell(row=stepCell.row, column=stepCell.col_idx+3).value]
			stepCount += 1 # Move to the next row
			
			stepCell = spreadsheet.cell(row=stepHeaderCell.row + stepCount, column=stepHeaderCell.col_idx) # Continue to the next step
		
		return stepData
	
	
	def getMetadata(self, spreadsheet):
		metadata = {}
		try:
			metadata["nameKey"] =           self.getAdjacentCellForKey(spreadsheet, self.config["nameKey"]).value # Get the text of the adjacent cell for a cell with a key
		except AttributeError:
			metadata["nameKey"] =           None
			
		try:
			metadata["descriptionKey"] =    self.getAdjacentCellForKey(spreadsheet, self.config["descriptionKey"]).value
		except AttributeError:
			metadata["descriptionKey"] =    None
			
		try:
			metadata["additionalInfoKey"] = self.getAdjacentCellForKey(spreadsheet, self.config["additionalInfoKey"]).value
		except AttributeError:
			metadata["additionalInfoKey"] = None
			
		try:
			metadata["priorityKey"] =       self.getAdjacentCellForKey(spreadsheet, self.config["priorityKey"]).value
		except AttributeError:
			metadata["priorityKey"] =       None
			
		try:
			metadata["authorKey"] =         self.getAdjacentCellForKey(spreadsheet, self.config["authorKey"]).value
		except AttributeError:
			metadata["authorKey"] =         None
			
		try:
			metadata["statusKey"] =         self.getAdjacentCellForKey(spreadsheet, self.config["statusKey"]).value
		except AttributeError:
			metadata["statusKey"] =         None
		
		return metadata
		
	
	def loadTest(self, spreadsheet): # Load a test from an excel spreadsheet, returns TestScript object
		metadata = self.getMetadata(spreadsheet)
		stepData = self.loadStepData(spreadsheet, self.getCellContaining(spreadsheet, self.config["stepsHeader"])) # Get step data
		
		# See if need to add additional information
		testDescription = metadata["descriptionKey"]
		if metadata["additionalInfoKey"] != None:
			testDescription += "\n\n --- Additional Information --- \n\n"
			testDescription += metadata["additionalInfoKey"]
			
		# Create TestScript Object for this spreadsheet
		testScript = TestScript(metadata["nameKey"], testDescription, metadata["priorityKey"], metadata["authorKey"], metadata["statusKey"])
		
		# Add Steps to the TestScript object
		for stepNumber in stepData:
			#                     number         description          expected result
			testScript.addStep(stepNumber, stepData[stepNumber][0], stepData[stepNumber][1])
		
		return testScript
	
	
	def loadSpreadsheet(self, filePath): # Load a spreadsheet from an excel file path
		workbook = load_workbook(filePath)
		if self.config["sheetName"] == "active":
			return workbook.active
		else:
			return workbook[self.config["sheetName"]]
	
	
	def load(self):	# Load the testscrips into the instance variable
		excelFiles = self.getFiles()
		doneCount = 1
		
		#print("Excel Files: " + str(excelFiles))
		for excelFile in excelFiles:
			print("Loading: " + str(excelFile) + ", %.2f%%" % (doneCount/len(excelFiles)*100))
			spreadsheet = self.loadSpreadsheet(excelFile)
			testScript = self.loadTest(spreadsheet)
			
			self.testScripts.append(testScript)
			doneCount += 1

		print("Loaded Scripts")
	
	